import csv
import hashlib
import io
import re

import requests
from openpyxl import load_workbook

UA = "WaveAlphaSourceQualification/1.0 (+public-safe-one-shot-probe)"
TIMEOUT = 20


def fetch(url):
    response = requests.get(url, headers={"User-Agent": UA, "Accept": "*/*"}, timeout=TIMEOUT, allow_redirects=True)
    print(f"HTTP url={url} status={response.status_code} type={response.headers.get('content-type')} bytes={len(response.content)} final={response.url}")
    print(f"SHA256 {hashlib.sha256(response.content).hexdigest()}")
    return response


def probe_arkb():
    print("\n=== ARKB CSV ===")
    url = "https://assets.ark-funds.com/fund-documents/funds-etf-csv/ARK_21SHARES_BITCOIN_ETF_ARKB_HOLDINGS.csv"
    r = fetch(url)
    if r.status_code != 200:
        print("ARKB_RESULT=NOT_PROVEN_HTTP")
        return
    text = r.content.decode("utf-8-sig", errors="replace")
    rows = list(csv.DictReader(io.StringIO(text)))
    print(f"ARKB_COLUMNS={list(rows[0].keys()) if rows else []}")
    print(f"ARKB_ROWS={len(rows)}")
    for i, row in enumerate(rows[:3]):
        compact = {k: v for k, v in row.items() if v not in (None, "")}
        print(f"ARKB_ROW_{i}={compact}")
    print("ARKB_RESULT=PASS" if rows else "ARKB_RESULT=NOT_PROVEN_EMPTY")


def probe_defi():
    print("\n=== DEFI XLSX ===")
    url = "https://hdx-website-cms-prod-upload-bucket.s3.amazonaws.com/DEFI_Holdings.xlsx"
    r = fetch(url)
    if r.status_code != 200:
        print("DEFI_RESULT=NOT_PROVEN_HTTP")
        return
    try:
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception as exc:
        print(f"DEFI_RESULT=NOT_PROVEN_WORKBOOK error={type(exc).__name__}:{exc}")
        return
    print(f"DEFI_SHEETS={wb.sheetnames}")
    for ws in wb.worksheets:
        print(f"DEFI_SHEET name={ws.title!r} rows={ws.max_row} cols={ws.max_column}")
        shown = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True):
            values = [v for v in row[:12]]
            if any(v not in (None, "") for v in values):
                print(f"DEFI_ROW_{shown}={values}")
                shown += 1
                if shown >= 15:
                    break
    print("DEFI_RESULT=PASS")


def probe_fidelity():
    print("\n=== FBTC FIDELITY ===")
    urls = {
        "summary": "https://digital.fidelity.com/prgw/digital/research/quote/dashboard/summary?symbol=FBTC",
        "basket": "https://research2.fidelity.com/fidelity/screeners/etf/etfholdings.asp?symbol=FBTC&view=Holdings",
    }
    for label, url in urls.items():
        r = fetch(url)
        text = r.text if r.status_code == 200 else ""
        normalized = re.sub(r"\s+", " ", text)
        print(
            f"FBTC_{label.upper()}_FIELDS="
            f"shares_outstanding={bool(re.search(r'Shares\s+outstanding', normalized, re.I))},"
            f"total_bitcoin_in_fund={bool(re.search(r'Total\s+bitcoin\s+in\s+fund', normalized, re.I))},"
            f"basket_holdings={bool(re.search(r'Basket\s+Holdings', normalized, re.I))},"
            f"bitcoin={bool(re.search(r'Bitcoin', normalized, re.I))},"
            f"nav={bool(re.search(r'\bNav\b', normalized, re.I))}"
        )
        date_match = re.search(r"AS OF\s+([0-9/\-]+)", normalized, re.I)
        if date_match:
            print(f"FBTC_{label.upper()}_ASOF={date_match.group(1)}")
    print("FBTC_RESULT=OBSERVED")


if __name__ == "__main__":
    probe_arkb()
    probe_defi()
    probe_fidelity()
