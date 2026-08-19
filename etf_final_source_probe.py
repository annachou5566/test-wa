import csv
import hashlib
import io
import re

import requests
from openpyxl import load_workbook

TIMEOUT = 25
UA = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/124 Safari/537.36 WaveAlphaPublicQA/1.0"


def fetch(label, url):
    try:
        r = requests.get(url, headers={"User-Agent": UA, "Accept": "*/*"}, timeout=TIMEOUT, allow_redirects=True)
        print(f"{label}_HTTP={r.status_code} type={r.headers.get('content-type')} bytes={len(r.content)} final={r.url}")
        print(f"{label}_SHA256={hashlib.sha256(r.content).hexdigest()}")
        return r
    except Exception as exc:
        print(f"{label}_FETCH_ERROR={type(exc).__name__}:{exc}")
        return None


def probe_arkb():
    print("=== ARKB_CURRENT_CSV ===")
    url = "https://assets.ark-funds.com/fund-documents/funds-etf-csv/ARK_21SHARES_BITCOIN_ETF_ARKB_HOLDINGS.csv"
    r = fetch("ARKB", url)
    if r is None or r.status_code != 200:
        print("ARKB_RESULT=NOT_PROVEN_HTTP")
        return
    text = r.content.decode("utf-8-sig", errors="replace")
    rows = list(csv.DictReader(io.StringIO(text)))
    print(f"ARKB_ROWS={len(rows)}")
    print(f"ARKB_COLUMNS={list(rows[0].keys()) if rows else []}")
    for i, row in enumerate(rows[:5]):
        compact = {k: v for k, v in row.items() if v not in (None, "")}
        print(f"ARKB_ROW_{i}={compact}")
    print("ARKB_RESULT=PASS" if rows else "ARKB_RESULT=NOT_PROVEN_EMPTY")


def probe_defi():
    print("=== DEFI_CURRENT_XLSX ===")
    url = "https://hdx-website-cms-prod-upload-bucket.s3.amazonaws.com/DEFI_Holdings.xlsx"
    r = fetch("DEFI", url)
    if r is None or r.status_code != 200:
        print("DEFI_RESULT=NOT_PROVEN_HTTP")
        return
    try:
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception as exc:
        print(f"DEFI_RESULT=NOT_PROVEN_WORKBOOK error={type(exc).__name__}:{exc}")
        return
    print(f"DEFI_SHEETS={wb.sheetnames}")
    for ws in wb.worksheets:
        print(f"DEFI_SHEET name={ws.title!r} rows={ws.max_row} cols={ws.max_column}")
        shown = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), values_only=True):
            vals = list(row[:14])
            if any(v not in (None, "") for v in vals):
                print(f"DEFI_ROW_{shown}={vals}")
                shown += 1
                if shown >= 20:
                    break
    print("DEFI_RESULT=PASS")


def marker_snippet(text, marker):
    m = re.search(marker, text, re.I)
    if not m:
        return None
    a = max(0, m.start() - 120)
    b = min(len(text), m.end() + 260)
    return re.sub(r"\s+", " ", text[a:b])


def probe_fbtc():
    print("=== FBTC_FIRST_PARTY ===")
    urls = {
        "SUMMARY": "https://digital.fidelity.com/prgw/digital/research/quote/dashboard/summary?symbol=FBTC",
        "HOLDINGS": "https://research2.fidelity.com/fidelity/screeners/etf/etfholdings.asp?symbol=FBTC&view=Holdings",
    }
    markers = {
        "SHARES": r"Shares\s+outstanding",
        "TOTAL_BTC": r"Total\s+bitcoin\s+in\s+fund",
        "BASKET": r"Basket\s+Holdings",
        "BITCOIN": r"Bitcoin",
        "NAV": r"\bNAV\b",
        "CREATION": r"Creation\s+(?:Unit|Basket)",
    }
    for label, url in urls.items():
        r = fetch(f"FBTC_{label}", url)
        if r is None:
            continue
        text = r.text if r.status_code == 200 else ""
        normalized = re.sub(r"\s+", " ", text)
        present = {name: bool(re.search(pattern, normalized, re.I)) for name, pattern in markers.items()}
        print(f"FBTC_{label}_MARKERS={present}")
        for name, pattern in markers.items():
            snippet = marker_snippet(normalized, pattern)
            if snippet:
                print(f"FBTC_{label}_{name}_SNIPPET={snippet[:500]}")
    print("FBTC_RESULT=OBSERVED")


if __name__ == "__main__":
    probe_arkb()
    probe_defi()
    probe_fbtc()
